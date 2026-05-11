# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"C:\nm-training\menus_macros.xlsx"

# Colors
BLUE_HEADER_BG   = "BDD7EE"
ALT_ROW_BG       = "F2F2F2"
SUBTOTAL_BG      = "E2EFDA"
TAB_BLUE         = "4472C4"
TAB_GREEN        = "70AD47"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():
    return Font(name="Arial", size=10, bold=True)

def regular_font():
    return Font(name="Arial", size=10)

def bold_font():
    return Font(name="Arial", size=10, bold=True)

def header_fill():
    return PatternFill("solid", fgColor=BLUE_HEADER_BG)

def alt_fill():
    return PatternFill("solid", fgColor=ALT_ROW_BG)

def subtotal_fill():
    return PatternFill("solid", fgColor=SUBTOTAL_BG)

def no_fill():
    return PatternFill(fill_type=None)

def apply_cell(cell, value, font=None, fill=None, alignment=None, border=None):
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border

# ─── WORKBOOK ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ─── RESUMEN ───────────────────────────────────────────────────────────────
ws_r = wb.active
ws_r.title = "Resumen"
ws_r.sheet_properties.tabColor = TAB_BLUE

resumen_headers = ["Menú", "kcal Total", "Proteína (g)", "Grasas (g)", "Carbos (g)"]
resumen_data = [
    ["🏋️ Gym 19h — Déficit", 2086, 164, 65, 206],
    ["🏋️ Gym 21h — Déficit", 1977, 151, 53, 218],
    ["⚽ Futbolito Jue 21h",  1898, 144, 37, 238],
    ["⚽ Fútbol Sáb 9:30h",   2086, 165, 36, 267],
    ["🏋️ Gym Dom 12h",        2046, 166, 65, 189],
    ["😴 Descanso Mié",       1851, 141, 60, 183],
    ["🍝 Viernes Carga Carbos", 2156, 164, 24, 317],
]

# Headers row 1
for ci, h in enumerate(resumen_headers, 1):
    cell = ws_r.cell(row=1, column=ci, value=h)
    cell.font      = header_font()
    cell.fill      = header_fill()
    cell.border    = thin_border()
    cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")

# Data rows
for ri, row in enumerate(resumen_data, 2):
    is_even = (ri % 2 == 0)
    fill = alt_fill() if is_even else no_fill()
    for ci, val in enumerate(row, 1):
        cell = ws_r.cell(row=ri, column=ci, value=val)
        cell.font   = regular_font()
        cell.fill   = fill
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="right" if ci > 1 else "left")

# Column widths
ws_r.column_dimensions["A"].width = 32
for col in ["B","C","D","E"]:
    ws_r.column_dimensions[col].width = 14

# Freeze row 1
ws_r.freeze_panes = "A2"

# ─── DETALLE ───────────────────────────────────────────────────────────────
ws_d = wb.create_sheet("Detalle")
ws_d.sheet_properties.tabColor = TAB_GREEN

detalle_headers = ["Menú", "Comida", "Alimento", "Cantidad (g)", "kcal", "Prot (g)", "Gras (g)", "Carb (g)"]

# Headers row 1
for ci, h in enumerate(detalle_headers, 1):
    cell = ws_d.cell(row=1, column=ci, value=h)
    cell.font      = header_font()
    cell.fill      = header_fill()
    cell.border    = thin_border()
    cell.alignment = Alignment(horizontal="right" if ci >= 4 else "left")

# Detalle rows: (row_num, is_subtotal, [values])
# Subtotal rows have values only in cols A, E-H (cols B-D empty)
detalle_rows = [
    # Gym 19h
    (2,  False, ["🏋️ Gym 19h", "desayuno",  "Huevos revueltos ×3",    180, 279, 23, 19, 2]),
    (3,  False, ["🏋️ Gym 19h", "desayuno",  "Pan integral ×2 taj",     60, 140,  6,  2, 25]),
    (4,  False, ["🏋️ Gym 19h", "desayuno",  "Café solo",              200,   5,  0,  0,  1]),
    (5,  False, ["🏋️ Gym 19h", "colacion1", "Manzana",                150,  78,  0,  0, 21]),
    (6,  False, ["🏋️ Gym 19h", "colacion1", "Almendras 30g",           30, 174,  6, 15,  5]),
    (7,  False, ["🏋️ Gym 19h", "almuerzo",  "Pechuga de pollo 150g",  150, 248, 47,  5,  0]),
    (8,  False, ["🏋️ Gym 19h", "almuerzo",  "Arroz cocido 180g",      180, 234,  5,  1, 51]),
    (9,  False, ["🏋️ Gym 19h", "almuerzo",  "Ensalada verde",         100,  25,  1,  0,  4]),
    (10, False, ["🏋️ Gym 19h", "colacion2", "Plátano pre-gym",        120, 107,  1,  0, 27]),
    (11, False, ["🏋️ Gym 19h", "colacion2", "Avena 40g pre-gym",       40, 156,  7,  3, 27]),
    (12, False, ["🏋️ Gym 19h", "cena",      "Scop Nico 30g (PW)",      30, 114, 22,  1,  3]),
    (13, False, ["🏋️ Gym 19h", "cena",      "Yogur Nestlé ×3 (PW)",   300, 340, 30,  6, 39]),
    (14, False, ["🏋️ Gym 19h", "cena",      "Huevos ×2 cena",         120, 186, 16, 13,  1]),
    (15, True,  ["SUBTOTAL Gym 19h", "", "", "", 2086, 164, 65, 206]),
    # Gym 21h
    (16, False, ["🏋️ Gym 21h", "desayuno",  "Huevos revueltos ×3",    180, 279, 23, 19, 2]),
    (17, False, ["🏋️ Gym 21h", "desayuno",  "Pan integral ×2 taj",     60, 140,  6,  2, 25]),
    (18, False, ["🏋️ Gym 21h", "desayuno",  "Café solo",              200,   5,  0,  0,  1]),
    (19, False, ["🏋️ Gym 21h", "colacion1", "Manzana",                150,  78,  0,  0, 21]),
    (20, False, ["🏋️ Gym 21h", "colacion1", "Almendras 30g",           30, 174,  6, 15,  5]),
    (21, False, ["🏋️ Gym 21h", "almuerzo",  "Pechuga de pollo 150g",  150, 248, 47,  5,  0]),
    (22, False, ["🏋️ Gym 21h", "almuerzo",  "Arroz cocido 180g",      180, 234,  5,  1, 51]),
    (23, False, ["🏋️ Gym 21h", "almuerzo",  "Ensalada verde",         100,  25,  1,  0,  4]),
    (24, False, ["🏋️ Gym 21h", "colacion2", "Avena 60g merienda",      60, 233, 10,  4, 40]),
    (25, False, ["🏋️ Gym 21h", "colacion2", "Plátano pre-gym 19:30",  120, 107,  1,  0, 27]),
    (26, False, ["🏋️ Gym 21h", "cena",      "Scop Nico 30g (PW)",      30, 114, 22,  1,  3]),
    (27, False, ["🏋️ Gym 21h", "cena",      "Yogur Nestlé ×3 (PW)",   300, 340, 30,  6, 39]),
    (28, True,  ["SUBTOTAL Gym 21h", "", "", "", 1977, 151, 53, 218]),
    # Futbolito Jue
    (29, False, ["⚽ Futbolito Jue", "desayuno",  "Huevos revueltos ×3",   180, 279, 23, 19, 2]),
    (30, False, ["⚽ Futbolito Jue", "desayuno",  "Pan integral ×2 taj",    60, 140,  6,  2, 25]),
    (31, False, ["⚽ Futbolito Jue", "desayuno",  "Café solo",             200,   5,  0,  0,  1]),
    (32, False, ["⚽ Futbolito Jue", "colacion1", "Manzana",               150,  78,  0,  0, 21]),
    (33, False, ["⚽ Futbolito Jue", "almuerzo",  "Pechuga de pollo 150g", 150, 248, 47,  5,  0]),
    (34, False, ["⚽ Futbolito Jue", "almuerzo",  "Arroz cocido 200g",     200, 260,  5,  1, 56]),
    (35, False, ["⚽ Futbolito Jue", "almuerzo",  "Ensalada verde",        100,  25,  1,  0,  4]),
    (36, False, ["⚽ Futbolito Jue", "colacion2", "Avena 50g + Plátano",   170, 302,  9,  3, 60]),
    (37, False, ["⚽ Futbolito Jue", "cena",      "Plátano pre-juego",     120, 107,  1,  0, 27]),
    (38, False, ["⚽ Futbolito Jue", "cena",      "Scop Nico 30g (PW)",     30, 114, 22,  1,  3]),
    (39, False, ["⚽ Futbolito Jue", "cena",      "Yogur Nestlé ×3 (PW)",  300, 340, 30,  6, 39]),
    (40, True,  ["SUBTOTAL Futbolito Jue", "", "", "", 1898, 144, 37, 238]),
    # Fútbol Sáb
    (41, False, ["⚽ Fútbol Sáb", "desayuno",  "Plátano ×2 pre-partido", 240, 214,  3,  1, 54]),
    (42, False, ["⚽ Fútbol Sáb", "desayuno",  "Pan integral ×3 taj",     90, 210,  9,  3, 38]),
    (43, False, ["⚽ Fútbol Sáb", "desayuno",  "Café solo",              200,   5,  0,  0,  1]),
    (44, False, ["⚽ Fútbol Sáb", "colacion1", "Scop Nico 30g (PW)",      30, 114, 22,  1,  3]),
    (45, False, ["⚽ Fútbol Sáb", "colacion1", "Yogur Nestlé ×3 (PW)",   300, 340, 30,  6, 39]),
    (46, False, ["⚽ Fútbol Sáb", "almuerzo",  "Pechuga de pollo 150g",  150, 248, 47,  5,  0]),
    (47, False, ["⚽ Fútbol Sáb", "almuerzo",  "Arroz cocido 200g",      200, 260,  5,  1, 56]),
    (48, False, ["⚽ Fútbol Sáb", "almuerzo",  "Ensalada verde",         100,  25,  1,  0,  4]),
    (49, False, ["⚽ Fútbol Sáb", "colacion2", "Manzana + Almendras",    180, 252,  6, 15, 26]),
    (50, False, ["⚽ Fútbol Sáb", "cena",      "Pechuga de pollo 120g",  120, 198, 37,  4,  0]),
    (51, False, ["⚽ Fútbol Sáb", "cena",      "Arroz cocido 150g",      150, 195,  4,  0, 42]),
    (52, False, ["⚽ Fútbol Sáb", "cena",      "Ensalada verde cena",    100,  25,  1,  0,  4]),
    (53, True,  ["SUBTOTAL Fútbol Sáb", "", "", "", 2086, 165, 36, 267]),
    # Gym Dom
    (54, False, ["🏋️ Gym Dom", "desayuno",  "Huevos revueltos ×3",   180, 279, 23, 19, 2]),
    (55, False, ["🏋️ Gym Dom", "desayuno",  "Pan integral ×2 taj",    60, 140,  6,  2, 25]),
    (56, False, ["🏋️ Gym Dom", "desayuno",  "Café solo",             200,   5,  0,  0,  1]),
    (57, False, ["🏋️ Gym Dom", "colacion1", "Avena 60g + Plátano",   180, 340, 11,  4, 67]),
    (58, False, ["🏋️ Gym Dom", "colacion1", "Almendras 30g",          30, 174,  6, 15,  5]),
    (59, False, ["🏋️ Gym Dom", "almuerzo",  "Scop Nico 30g (PW)",     30, 114, 22,  1,  3]),
    (60, False, ["🏋️ Gym Dom", "almuerzo",  "Yogur Nestlé ×3 (PW)", 300, 340, 30,  6, 39]),
    (61, False, ["🏋️ Gym Dom", "colacion2", "Pechuga de pollo 150g", 150, 248, 47,  5,  0]),
    (62, False, ["🏋️ Gym Dom", "colacion2", "Arroz cocido 150g",     150, 195,  4,  0, 42]),
    (63, False, ["🏋️ Gym Dom", "cena",      "Ensalada + Huevos ×2",  220, 211, 17, 13,  5]),
    (64, True,  ["SUBTOTAL Gym Dom", "", "", "", 2046, 166, 65, 189]),
    # Descanso Mié
    (65, False, ["😴 Descanso Mié", "desayuno",  "Huevos revueltos ×3",   180, 279, 23, 19, 2]),
    (66, False, ["😴 Descanso Mié", "desayuno",  "Pan integral ×2 taj",    60, 140,  6,  2, 25]),
    (67, False, ["😴 Descanso Mié", "desayuno",  "Café solo",             200,   5,  0,  0,  1]),
    (68, False, ["😴 Descanso Mié", "colacion1", "Manzana",               150,  78,  0,  0, 21]),
    (69, False, ["😴 Descanso Mié", "colacion1", "Almendras 30g",          30, 174,  6, 15,  5]),
    (70, False, ["😴 Descanso Mié", "almuerzo",  "Pechuga de pollo 150g", 150, 248, 47,  5,  0]),
    (71, False, ["😴 Descanso Mié", "almuerzo",  "Arroz cocido 180g",     180, 234,  5,  1, 51]),
    (72, False, ["😴 Descanso Mié", "almuerzo",  "Ensalada verde",        100,  25,  1,  0,  4]),
    (73, False, ["😴 Descanso Mié", "colacion2", "Avena 50g + Plátano",   170, 302,  9,  3, 60]),
    (74, False, ["😴 Descanso Mié", "cena",      "Atún al agua 100g",     100, 110, 25,  1,  0]),
    (75, False, ["😴 Descanso Mié", "cena",      "Huevos ×2",             120, 186, 16, 13,  1]),
    (76, False, ["😴 Descanso Mié", "cena",      "Pan integral ×1 taj",    30,  70,  3,  1, 13]),
    (77, True,  ["SUBTOTAL Descanso Mié", "", "", "", 1851, 141, 60, 183]),
    # Viernes Carbos
    (78, False, ["🍝 Viernes Carbos", "desayuno",  "Avena 70g + Plátano",    190, 379, 13,  5, 73]),
    (79, False, ["🍝 Viernes Carbos", "desayuno",  "Café solo",              200,   5,  0,  0,  1]),
    (80, False, ["🍝 Viernes Carbos", "colacion1", "Pan ×2 + Fruta",         210, 218,  6,  2, 46]),
    (81, False, ["🍝 Viernes Carbos", "almuerzo",  "Pechuga de pollo 150g",  150, 248, 47,  5,  0]),
    (82, False, ["🍝 Viernes Carbos", "almuerzo",  "Fideos cocidos 200g",    200, 316, 12,  2, 62]),
    (83, False, ["🍝 Viernes Carbos", "almuerzo",  "Ensalada verde",         100,  25,  1,  0,  4]),
    (84, False, ["🍝 Viernes Carbos", "colacion2", "Plátano pre-gym",        120, 107,  1,  0, 27]),
    (85, False, ["🍝 Viernes Carbos", "cena",      "Scop Nico 30g (PW)",      30, 114, 22,  1,  3]),
    (86, False, ["🍝 Viernes Carbos", "cena",      "Yogur Nestlé ×3 (PW)",   300, 340, 30,  6, 39]),
    (87, False, ["🍝 Viernes Carbos", "cena",      "Fideos con atún",         280, 404, 32,  3, 62]),
    (88, True,  ["SUBTOTAL Viernes Carbos", "", "", "", 2156, 164, 24, 317]),
]

subtotal_row_nums = {r[0] for r in detalle_rows if r[1]}

for row_num, is_subtotal, values in detalle_rows:
    if is_subtotal:
        # subtotal: bold, green bg
        f    = bold_font()
        fill = subtotal_fill()
        # values layout: [label, "", "", "", kcal, prot, gras, carb]
        for ci, val in enumerate(values, 1):
            cell = ws_d.cell(row=row_num, column=ci, value=val if val != "" else None)
            cell.font      = f
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="right" if ci >= 4 else "left")
    else:
        # determine alternate shading: shade odd data rows (rows 3,5,7... excluding subtotals)
        # "non-subtotal odd rows" – check if row_num is odd
        is_odd = (row_num % 2 != 0)
        fill = alt_fill() if is_odd else no_fill()
        for ci, val in enumerate(values, 1):
            cell = ws_d.cell(row=row_num, column=ci, value=val)
            cell.font      = regular_font()
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="right" if ci >= 4 else "left")

# Column widths Detalle
col_widths_d = [28, 16, 30, 14, 10, 10, 10, 10]
for i, w in enumerate(col_widths_d, 1):
    ws_d.column_dimensions[get_column_letter(i)].width = w

# Freeze row 1
ws_d.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")

import os
print("File exists:", os.path.exists(OUTPUT))
print("File size:", os.path.getsize(OUTPUT), "bytes")
